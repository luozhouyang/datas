from openpyxl import Workbook
import os


class XLSXSaver:

    def __init__(self, filename, ages, services, educations, origins, cities):
        self.filename = os.path.join(os.path.dirname(__file__), "xlsx", filename)
        self.ages_all_dict = ages[0]
        self.ages_men_dict = ages[1]
        self.ages_women_dict = ages[2]
        self.services_all_dict = services[0]
        self.services_men_dict = services[1]
        self.services_women_dict = services[2]
        self.edus_all_dict = educations[0]
        self.edus_men_dict = educations[1]
        self.edus_women_dict = educations[2]
        self.origins_all_dict = origins[0]
        self.origins_men_dict = origins[1]
        self.origins_women_dict = origins[2]
        self.cities_all_dict = cities[0]
        self.cities_men_dict = cities[1]
        self.cities_women_dict = cities[2]

    @staticmethod
    def _save_xlsx_horizontal(wb, title, index, c0, c1, c2):
        sheet = wb.create_sheet(title=title, index=index)
        row = ['']
        men_row = ['男']
        women_row = ['女']
        for k, _ in c0.items():
            row.append(k)
            if k in c1.keys():
                v0 = int(c1[k])
            else:
                v0 = 0
            men_row.append(v0)
            if k in c2.keys():
                v1 = int(c2[k])
            else:
                v1 = 0
            women_row.append(v1)
        sheet.append(row)
        sheet.append(men_row)
        sheet.append(women_row)

    @staticmethod
    def _save_xlsx_vertical(wb, title, index, c0, c1, c2):
        sheet = wb.create_sheet(title, index)
        row = ['', '男', '女']
        sheet.append(row)
        for k, _ in c0.items():
            if k in c1.keys():
                v_men = int(c1[k])
            else:
                v_men = 0
            if k in c2.keys():
                v_women = int(c2[k])
            else:
                v_women = 0
            row = [k, v_men, v_women]
            sheet.append(row)

    def _save_xlsx(self, wb, title, index, c0, c1, c2):
        self._save_xlsx_horizontal(wb, title + "_1", index, c0, c1, c2)
        self._save_xlsx_vertical(wb, title + "_2", index + 5, c0, c1, c2)

    def _save_age_xlsx(self, wb):
        self._save_xlsx(wb, title='age', index=0,
                        c0=self.ages_all_dict, c1=self.ages_men_dict, c2=self.ages_women_dict)

    def _save_service_type_xlsx(self, wb):
        self._save_xlsx(wb, title='service_type', index=1,
                        c0=self.services_all_dict, c1=self.services_men_dict, c2=self.services_women_dict)

    def _save_edu_xlsx(self, wb):
        self._save_xlsx(wb, title='education', index=2,
                        c0=self.edus_all_dict, c1=self.edus_men_dict, c2=self.edus_women_dict)

    def _save_origin_xlsx(self, wb):
        self._save_xlsx(wb, title='origin', index=3,
                        c0=self.origins_all_dict, c1=self.origins_men_dict, c2=self.origins_women_dict)

    def _save_lives_xlsx(self, wb):
        self._save_xlsx(wb, title='lives_in', index=4,
                        c0=self.cities_all_dict, c1=self.cities_men_dict, c2=self.cities_women_dict)

    def save_to_xlsx(self):
        wb = Workbook()
        self._save_age_xlsx(wb)
        self._save_service_type_xlsx(wb)
        self._save_edu_xlsx(wb)
        self._save_origin_xlsx(wb)
        self._save_lives_xlsx(wb)
        wb.save(self.filename)
